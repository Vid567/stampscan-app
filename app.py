"""
StampScan 2.0 - Gradio app (Render / Hugging Face Space)
Upload a photo -> Roboflow Serverless Workflow -> numbered preview + table + Excel
The Roboflow API key is read from the ROBOFLOW_API_KEY environment variable
(never in the browser). It must be the workspace PRIVATE key -- the publishable
"rf_..." key is only valid for the client-side inferencejs library and will be
rejected by the serverless endpoint.
"""

import os
import re
import json
import base64
import traceback
from datetime import datetime
from pathlib import Path

import gradio as gr
from inference_sdk import InferenceHTTPClient

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ---- Config -----------------------------------------------------------------
API_KEY = os.environ.get("ROBOFLOW_API_KEY", "")
WORKSPACE = "vidcas567-gmail-com"
WORKFLOW_ID = "automated-stamp-scanner-1787234733118"
API_URL = "https://serverless.roboflow.com"

client = InferenceHTTPClient(api_url=API_URL, api_key=API_KEY) if API_KEY else None

# Gradio 4 will only serve files from directories it has been told about.
# Anything left in the system temp dir can be built fine and still refuse to
# download, so outputs go in a folder next to the app and that folder is
# passed to launch(allowed_paths=...).
OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

COLUMNS = [
    "#",
    "Country",
    "Issue / Series",
    "Year",
    "Denomination",
    "Catalog match",
    "OCR text",
    "Condition",
    "Confidence",
]


# ---- Response normalisation --------------------------------------------------
# Roboflow workflow blocks are inconsistent: a block may hand back a plain list,
# or a dict that wraps the list under "predictions", or a single dict for one
# item. Everything downstream wants a list, so funnel all of it through here.
def _as_list(value):
    """Coerce a workflow output value into a list of entries."""
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        # Object-detection blocks return {"predictions": [...], "image": {...}}
        for key in ("predictions", "results", "items", "data"):
            nested = value.get(key)
            if isinstance(nested, list):
                return nested
        # A dict keyed "0", "1", "2"... is a list in disguise
        if value and all(str(k).isdigit() for k in value.keys()):
            return [value[k] for k in sorted(value.keys(), key=lambda k: int(k))]
        # Otherwise it is a single entry
        return [value]
    return [value]


def _first_output(result):
    """The SDK returns a list of step outputs; older builds return a bare dict."""
    if isinstance(result, list):
        return result[0] if result else {}
    if isinstance(result, dict):
        return result
    return {}


def _as_count(value, fallback):
    if isinstance(value, bool):
        return fallback
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    if isinstance(value, list):
        return len(value)
    return fallback


# ---- Field extraction --------------------------------------------------------
# The OCR and analysis blocks are language-model steps, so they return JSON as a
# string -- often wrapped in a markdown ```json fence. Dumping that raw into a
# spreadsheet cell is useless, so unwrap it and split it into real columns.
_FENCE = re.compile(r"^\s*```(?:json)?\s*|\s*```\s*$", re.IGNORECASE)


def _coerce(entry):
    """Get a dict out of a dict, a JSON string, or a fenced JSON string."""
    if isinstance(entry, dict):
        return entry
    if isinstance(entry, list):
        return _coerce(entry[0]) if entry else {}
    if isinstance(entry, str):
        raw = _FENCE.sub("", entry).strip()
        if raw.startswith("{") or raw.startswith("["):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    parsed = parsed[0] if parsed else {}
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                pass
        return {"all_text": entry}
    return {}


def _pick(d, *keys):
    """First non-empty value among keys, flattened to a string."""
    for k in keys:
        v = d.get(k)
        if v not in (None, "", [], {}):
            if isinstance(v, (dict, list)):
                return json.dumps(v, ensure_ascii=False)
            return str(v)
    return ""


def _condition_of(analysis):
    """Prefer an explicit condition field; otherwise assemble one."""
    explicit = _pick(analysis, "condition", "grade", "state")
    if explicit:
        return explicit
    parts = []
    for k in ("damage", "cancellation", "centering", "perforations", "gum"):
        v = analysis.get(k)
        if v not in (None, "", [], {}):
            parts.append(f"{k}: {v}")
    return ", ".join(parts)


# ---- Helpers ----------------------------------------------------------------
def _stamp():
    """Timestamp for output filenames, so repeat scans don't overwrite."""
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _decode_output_image(value):
    """Roboflow may return the annotated image as base64, a data URL, or a dict."""
    if not value:
        return None
    try:
        if isinstance(value, list):
            value = value[0] if value else ""
        if isinstance(value, dict):
            value = value.get("value") or value.get("image") or value.get("base64") or ""
        if not isinstance(value, str):
            return None
        if value.startswith("data:image"):
            value = value.split(",", 1)[1]
        raw = base64.b64decode(value)
        path = OUTPUT_DIR / f"stamp-preview-{_stamp()}.png"
        path.write_bytes(raw)
        return str(path)
    except Exception:
        return None


def _build_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stamps"
    ws.append(COLUMNS)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    widths = {"#": 5, "Country": 16, "Issue / Series": 46, "Year": 8,
              "Denomination": 14, "Catalog match": 20, "OCR text": 30,
              "Condition": 26, "Confidence": 11}
    for idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(name, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    path = OUTPUT_DIR / f"stamp-inventory-{_stamp()}.xlsx"
    wb.save(path)
    return str(path)


def _describe(value):
    """Compact description of a value's shape, for the server log."""
    if isinstance(value, dict):
        return f"dict(keys={list(value.keys())[:8]})"
    if isinstance(value, list):
        inner = _describe(value[0]) if value else "empty"
        return f"list(len={len(value)}, first={inner})"
    return type(value).__name__


# ---- Main scan --------------------------------------------------------------
def scan(image_path):
    if not API_KEY or client is None:
        return None, [], "⚠️ ROBOFLOW_API_KEY is not set for this service.", None
    if not image_path:
        return None, [], "Please upload a photo first.", None

    try:
        result = client.run_workflow(
            workspace_name=WORKSPACE,
            workflow_id=WORKFLOW_ID,
            images={"image": image_path},
            use_cache=True,
        )
    except Exception as e:
        return None, [], f"Error calling the workflow: {e}", None

    # Everything below is parsing. Never let it crash the UI -- a traceback in
    # the Render log is useful, red "Error" boxes are not.
    try:
        scan_out = _first_output(result)
        print("[stampscan] workflow output shape:", _describe(scan_out), flush=True)

        predictions = _as_list(scan_out.get("predictions"))
        ocr_results = _as_list(scan_out.get("ocr_results"))
        analysis = _as_list(scan_out.get("stamp_analysis"))

        n = max(len(predictions), len(ocr_results), len(analysis))
        count = _as_count(scan_out.get("stamp_count"), n)

        rows = []
        for i in range(n):
            pred = predictions[i] if i < len(predictions) else {}
            ocr = _coerce(ocr_results[i] if i < len(ocr_results) else {})
            ana = _coerce(analysis[i] if i < len(analysis) else {})

            conf = pred.get("confidence", "") if isinstance(pred, dict) else ""
            if isinstance(conf, (int, float)) and not isinstance(conf, bool):
                conf = f"{conf:.2f}"

            country = _pick(ana, "likely_country", "country") or _pick(ocr, "country_text", "country")
            denom = _pick(ana, "denomination", "value", "face_value") or _pick(ocr, "value_text")
            ocr_text = _pick(ocr, "all_text", "text", "ocr", "value_text")

            rows.append([
                i + 1,
                country,
                _pick(ana, "likely_issue", "issue", "series", "description"),
                _pick(ana, "likely_year", "year", "date"),
                denom,
                _pick(ana, "catalog_match", "catalog", "catalogue", "scott"),
                ocr_text.replace("\n", " ").strip(),
                _condition_of(ana),
                conf,
            ])

        annotated = _decode_output_image(scan_out.get("output_image"))
        xlsx = _build_excel(rows) if rows else None

        if n == 0:
            keys = list(scan_out.keys()) if isinstance(scan_out, dict) else []
            summary = f"The workflow ran but returned no stamps. Output keys were: {keys}"
        else:
            summary = (
                f"Found {count} stamp(s). "
                "Click the blue file size to the right of the filename to download the Excel."
            )

        return annotated, rows, summary, xlsx

    except Exception as e:
        traceback.print_exc()
        try:
            shape = _describe(_first_output(result))
        except Exception:
            shape = "unreadable"
        return (
            None,
            [],
            f"Could not read the workflow response ({type(e).__name__}: {e}). "
            f"Response shape was: {shape}",
            None,
        )


# ---- UI ---------------------------------------------------------------------
demo = gr.Interface(
    fn=scan,
    inputs=gr.Image(type="filepath", label="Upload an album-page photo"),
    outputs=[
        gr.Image(label="Detected stamps (numbered)"),
        gr.Dataframe(headers=COLUMNS, label="Results (review before trusting)", wrap=True),
        gr.Textbox(label="Summary"),
        gr.File(label="Download Excel"),
    ],
    title="StampScan 2.0",
    description=(
        "Upload a photo of a stamp album page. It detects each stamp, reads the "
        "printed text, suggests an identification, and checks visible condition. "
        "Counts on very dense pages are still approximate — always review the numbered preview."
    ),
    allow_flagging="never",
)

if __name__ == "__main__":
    demo.launch(
        server_name="0.0.0.0",
        server_port=int(os.environ.get("PORT", 7860)),
        allowed_paths=[str(OUTPUT_DIR.resolve())],
    )
